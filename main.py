from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, NamedStyle

from constant import *
from line_data_process import line_value_calibrate, line_value_modify, write_line_data


def write_initialize():
    """
    创建工作表，建立工作簿，写入第一行列名，添加日期样式，并返回该工作簿
    :return: 工作表和工作簿
    """
    # 输出文件：打开要写的Excel文件：处理数据.xlsx
    wb_write = Workbook()
    # 建立工作簿
    ws_write = wb_write.active
    ws_write.title = OUTPUT_FILE_SHEET_NAME
    # 写第一行所有列名
    for i, std_col_name in enumerate(COL_NAME_LIST):
        ws_write.cell(row=1, column=i + 1, value=std_col_name)
    # 定义一个日期样式，并添加到工作簿
    date_style = NamedStyle(name="date_style", number_format="YYYY/MM/DD")
    if "date_style" not in wb_write.style_names:
        wb_write.add_named_style(date_style)
    return wb_write, ws_write


def get_index_col_map(row, inverted_col_name_map):
    """
    获取有效列名和列序号的映射
    :param row: 行
    :param inverted_col_name_map: 原始文件的工作簿中列名到标准列名的映射
    :return: 有效列名和列序号的映射
    """
    # 有效列名和列序号的映射
    index_col_map = {}
    for col_index, col_name in enumerate(row):
        # 如果列名为有效真值（在预置的列名表中）
        if col_name in inverted_col_name_map.keys():
            # 有效列名和列序号的映射
            index_col_map[col_index] = inverted_col_name_map[col_name]
    return index_col_map


def get_line_data(row, index_col_map, sheet_name, row_index):
    """
    读取一行内所有有效信息
    :param row: 行
    :param index_col_map: 有效列名和列序号的映射
    :param sheet_name: 工作簿名
    :param row_index: 行号
    :return: 行内所有有效信息
    """
    line = []
    for col_index, value in enumerate(row):
        # 如果需要这一列数据
        if col_index in index_col_map.keys():
            # 列名
            col_name = index_col_map[col_index]
            # 根据列名检查数值
            flag, value = line_value_calibrate(col_name, value)
            # 不合格的行删除
            if not flag:
                print(f"删除: 表[{sheet_name}] 第{row_index}行 {col_name}={value}")
                line = None
                break
            # 将列名和数值的pair存入line中
            line.append({col_name: value})
    return line


def sheet_process(ref, wb_read, ws_write, write_row_index):
    """
    对原始文件中每一个工作簿分析处理，写入ws_write
    :param ref: 预置的参考信息
    :param wb_read: 原始文件中的工作簿
    :param ws_write: 要写的工作簿
    :param write_row_index: 写的工作簿的行号
    :return: 写的工作簿的行号
    """
    # 原始文件工作簿名
    sheet_name = ref['name']
    # 标准列名到原始文件的工作簿中列名的映射
    col_name_map = ref['col_name_map']
    # 原始文件的工作簿中列名到标准列名的映射
    inverted_col_name_map = {v: k for k, v in col_name_map.items()}
    # 获取原始文件的工作簿
    ws_read = wb_read[sheet_name]
    # 赋初始值
    index_col_map = {}
    # 遍历行读取数据
    for row_index, row in enumerate(ws_read.iter_rows(values_only=True)):
        # 跳过全为None的行
        if all(cell is None for cell in row):
            continue
        # 若为第一行，获取有效列名和列序号的映射
        if row_index == 0:
            index_col_map = get_index_col_map(row, inverted_col_name_map)
        # 非第一行，为数据行
        else:
            # 获取数据行所有有效数据
            line = get_line_data(row, index_col_map, sheet_name, row_index)
            if line is not None:
                # 按照预设逻辑，调整行内数值
                line = line_value_modify(line)
                # 写入ws_write，同步进行后处理
                ws_write = write_line_data(write_row_index, ws_write, line)
                write_row_index += 1
    return write_row_index


def save_workbook(wb_write, ws_write):
    """
    保存到Excel文件
    :param wb_write: 写入的工作表
    :param ws_write: 写入的工作簿
    """
    # 定义居中对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')
    # 遍历工作表中的所有单元格并应用居中对齐
    for row in ws_write.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
    wb_write.save(OUTPUT_FILE_PATH)


def main():
    # 要写的工作表和工作簿初始化
    wb_write, ws_write = write_initialize()
    # 加载原始文件
    wb_read = load_workbook(INPUT_FILE_PATH, data_only=True)
    write_row_index = 2
    # 遍历原始文件中每一个工作簿
    for ref in REF_LIST:
        write_row_index = sheet_process(ref, wb_read, ws_write, write_row_index)
    # 保存到Excel文件
    save_workbook(wb_write, ws_write)


if __name__ == "__main__":
    main()
